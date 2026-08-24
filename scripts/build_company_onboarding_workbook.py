"""Genera y valida el libro de incorporación empresarial SmartDiag504."""
from __future__ import annotations
import argparse
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SHEETS = {
 "Empresa": ["codigo_empresa","razon_social","nombre_comercial","rtn","direccion","telefono","correo","moneda","zona_horaria"],
 "Sucursales": ["codigo_sucursal","nombre","direccion","telefono","responsable","activa"],
 "Bodegas": ["codigo_bodega","nombre","codigo_sucursal","tipo","ubicacion","responsable","activa"],
 "Empleados": ["nombre_completo","identidad","fecha_nacimiento","direccion","telefono","correo_personal","numero_seguro_social","proveedor_seguro","puesto","tipo_contrato","fecha_inicio","fecha_fin","tipo_pago","salario_base_hnl","horas_semanales","sucursal","rol_sistema","activo"],
 "Asistencia_inicial": ["identidad_empleado","fecha","hora_entrada","hora_salida","horas_regulares","horas_extra","estado","observacion"],
 "Proveedores": ["codigo","nombre","rtn","correo","telefono","dias_credito","moneda","direccion","contacto","activo"],
 "Saldos_inventario": ["codigo_repuesto","codigo_bodega","cantidad","costo_unitario_hnl","lote","serie","fecha_corte","observacion"],
 "Clientes": ["codigo_cliente","nombre","identidad_rtn","telefono","correo","direccion","tipo_cliente","credito_solicitado","consentimiento_contacto"],
 "Vehiculos": ["vin","placa","codigo_cliente","marca","modelo","anio","motor","color","kilometraje","fecha_ultimo_aceite","km_proximo_aceite"],
 "Cotizaciones_abiertas": ["numero_origen","fecha","codigo_cliente","vin","vigencia_dias","moneda","impuesto_hnl","descuento_hnl","estado","observacion"],
 "Cotizacion_lineas": ["numero_origen","tipo_linea","codigo_item","descripcion","cantidad","precio_unitario_hnl","aprobada_cliente"],
 "Compras_abiertas": ["numero_origen","codigo_proveedor","codigo_sucursal","fecha","fecha_esperada","moneda","tasa_cambio","impuesto","estado","observacion"],
 "Compra_lineas": ["numero_origen","codigo_item","descripcion","cantidad","costo_unitario"],
 "Importaciones": ["numero_compra","incoterm","pais_origen","puerto_destino","eta","metodo_distribucion","flete","seguro","aduana","impuestos","manejo","otros"],
 "Empresas_envio": ["codigo","nombre","rtn","contacto","telefono","correo","cobertura","tarifa_base_hnl","tiempo_estimado","requiere_guia","activo"],
 "Envios_abiertos": ["numero_pedido","empresa_envio","numero_guia","destinatario","telefono","direccion","estado","fecha_envio","costo_hnl","url_evidencia"],
 "Usuarios_roles": ["identidad_empleado","correo_usuario","rol","sucursal","requiere_caja","requiere_erp","requiere_portal_movil","aprobado_por"],
 "Fiscalidad": ["rtn","cai","rango_desde","rango_hasta","fecha_limite_emision","tipo_documento","modo_impresion","impresora","aprobado_contador","observacion"],
 "Documentos": ["tipo_documento","nombre_archivo_referencia","tamano_papel","orientacion","logo","pie_pagina","campos_obligatorios","requiere_preimpreso","aprobado_por"],
 "Saldos_contables": ["fecha_corte","cuenta_contable","nombre_cuenta","debito_hnl","credito_hnl","centro_costo","sucursal","aprobado_contador"],
}

VALIDATIONS = [
 ("Empleados","J",'"PERMANENTE,TEMPORAL,POR_HORA,CONTRATISTA"'),
 ("Empleados","M",'"MENSUAL,QUINCENAL,SEMANAL,DIARIO,POR_HORA"'),
 ("Empleados","Q",'"ADMINISTRADOR,GERENCIA,CONTADOR,ASESOR,TECNICO,CAJA,BODEGA,RRHH,MARKETING"'),
 ("Asistencia_inicial","G",'"PRESENTE,AUSENTE,PERMISO,FERIADO"'),
 ("Cotizacion_lineas","B",'"REPUESTO,MANO_DE_OBRA,OTRO"'),
 ("Importaciones","F",'"POR_VALOR,POR_CANTIDAD,POR_PESO"'),
 ("Fiscalidad","G",'"PREIMPRESA,AUTOIMPRESOR,TERMICA,CARTA"'),
]

def build(path: Path) -> None:
 wb = Workbook(); info = wb.active; info.title = "LEEME"
 for row in [
  ["PAQUETE SMARTDIAG504","Versión 1.0"],
  ["Propósito","Recopilar y validar datos antes de escribir en ERPNext."],
  ["Regla","No cambie hojas ni encabezados. Una fila es un registro."],
  ["Seguridad","No incluya contraseñas, tokens ni claves bancarias."],
  ["Formato","Fechas AAAA-MM-DD, horas HH:MM, importes sin símbolos."],
  ["Catálogo","Use el archivo separado 01_catalogo_repuestos_mano_obra.xlsx."],
  ["Carga","Vista previa, corrección, aprobación y aplicación."],
 ]: info.append(row)
 for name, headers in SHEETS.items():
  ws=wb.create_sheet(name); ws.append(headers); ws.freeze_panes="A2"
  ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}1"
  for cell in ws[1]:
   cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="0B2A4A"); cell.alignment=Alignment(wrap_text=True)
  for i, header in enumerate(headers,1): ws.column_dimensions[get_column_letter(i)].width=min(max(len(header)+4,15),32)
 for sheet,col,formula in VALIDATIONS:
  dv=DataValidation(type="list",formula1=formula,allow_blank=True); wb[sheet].add_data_validation(dv); dv.add(f"{col}2:{col}5000")
 info.column_dimensions["A"].width=24; info.column_dimensions["B"].width=100
 for c in info[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="C8102E")
 path.parent.mkdir(parents=True,exist_ok=True); wb.save(path)

def validate(path: Path) -> None:
 wb=load_workbook(path,read_only=True,data_only=False)
 if wb.sheetnames != ["LEEME",*SHEETS]: raise SystemExit("Estructura de hojas inválida")
 for name,headers in SHEETS.items():
  actual=[c.value for c in next(wb[name].iter_rows(min_row=1,max_row=1))]
  if actual != headers: raise SystemExit(f"Encabezados inválidos: {name}")
 print(f"OK: {path} ({len(wb.sheetnames)} hojas)")

if __name__ == "__main__":
 p=argparse.ArgumentParser(); p.add_argument("output",type=Path); p.add_argument("--validate",action="store_true"); a=p.parse_args()
 if a.validate: validate(a.output)
 else: build(a.output); validate(a.output)
