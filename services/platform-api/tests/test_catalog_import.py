from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.catalog_import import build_catalog_workbook, parse_catalog_workbook


def test_template_contains_instructions_labor_and_parts_sheets() -> None:
    workbook = load_workbook(BytesIO(build_catalog_workbook()), data_only=True)

    assert workbook.sheetnames == ["Instrucciones", "Mano de obra", "Repuestos"]
    assert [cell.value for cell in workbook["Mano de obra"][1]] == [
        "codigo",
        "descripcion",
        "marca_vehiculo",
        "modelo_vehiculo",
        "anio_desde",
        "anio_hasta",
        "motor",
        "tiempo_horas",
        "precio_costo_hnl",
        "precio_venta_hnl",
        "activo",
    ]
    assert "tiempo_horas" not in [cell.value for cell in workbook["Repuestos"][1]]


def test_parser_groups_repeated_codes_as_vehicle_fitments() -> None:
    workbook = load_workbook(BytesIO(build_catalog_workbook()))
    labor = workbook["Mano de obra"]
    labor.append(
        [
            "MO-FRENOS",
            "Cambio de pastillas",
            "Toyota",
            "Corolla",
            2015,
            2020,
            "1.8",
            1.5,
            350,
            650,
            "SI",
        ]
    )
    labor.append(
        [
            "MO-FRENOS",
            "Cambio de pastillas",
            "Honda",
            "Civic",
            2016,
            2021,
            "2.0",
            1.5,
            350,
            650,
            "SI",
        ]
    )
    parts = workbook["Repuestos"]
    parts.append(
        [
            "REP-001",
            "Pastilla de freno",
            "04465-02190",
            "Akebono",
            "Unidad",
            "Toyota",
            "Corolla",
            2015,
            2020,
            "1.8",
            900,
            1350,
            "SI",
        ]
    )
    buffer = BytesIO()
    workbook.save(buffer)

    preview = parse_catalog_workbook(buffer.getvalue())

    assert preview.errors == []
    assert len(preview.labor) == 1
    assert preview.labor[0].code == "MO-FRENOS"
    assert len(preview.labor[0].fitments) == 2
    assert preview.parts[0].oem_number == "04465-02190"
    assert preview.parts[0].sale_price == 1350


def test_parser_reports_row_and_column_for_invalid_prices() -> None:
    workbook = load_workbook(BytesIO(build_catalog_workbook()))
    workbook["Mano de obra"].append(
        ["MO-001", "Diagnostico", "Toyota", "Corolla", 2010, 2020, "", 1, 800, 700, "SI"]
    )
    buffer = BytesIO()
    workbook.save(buffer)

    preview = parse_catalog_workbook(buffer.getvalue())

    assert preview.labor == []
    assert preview.errors[0].sheet == "Mano de obra"
    assert preview.errors[0].row == 2
    assert preview.errors[0].column == "precio_venta_hnl"
    assert "menor" in preview.errors[0].message
