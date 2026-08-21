from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

LABOR_HEADERS = [
    "codigo",
    "descripcion",
    "marca_vehiculo",
    "modelo_vehiculo",
    "anio_desde",
    "anio_hasta",
    "motor",
    "tiempo_horas",
    "precio_costo_hnl",
    "precio_venta_hnl",
    "activo",
]
PART_HEADERS = [
    "codigo",
    "descripcion",
    "numero_oem",
    "marca_repuesto",
    "unidad",
    "marca_vehiculo",
    "modelo_vehiculo",
    "anio_desde",
    "anio_hasta",
    "motor",
    "precio_costo_hnl",
    "precio_venta_hnl",
    "activo",
]


@dataclass(frozen=True, slots=True)
class VehicleFitment:
    make: str
    model: str
    year_from: int
    year_to: int
    engine: str | None = None


@dataclass(slots=True)
class CatalogEntry:
    code: str
    description: str
    cost_price: Decimal
    sale_price: Decimal
    active: bool
    fitments: list[VehicleFitment] = field(default_factory=list)
    standard_hours: Decimal | None = None
    oem_number: str | None = None
    brand: str | None = None
    unit: str | None = None


@dataclass(frozen=True, slots=True)
class ImportErrorRow:
    sheet: str
    row: int
    column: str
    message: str


@dataclass(slots=True)
class CatalogPreview:
    labor: list[CatalogEntry] = field(default_factory=list)
    parts: list[CatalogEntry] = field(default_factory=list)
    errors: list[ImportErrorRow] = field(default_factory=list)


DEMO_VEHICLES = [
    ("Ford", "Escape", 2020, 2020, "1.5L EcoBoost"),
    ("Ford", "F-150", 2020, 2020, "3.5L EcoBoost"),
    ("Honda", "Civic", 2008, 2008, "1.8L"),
]

DEMO_LABOR = [
    ("MO-DIAG-001", "Diagnostico electronico integral", 1, 380, 750),
    ("MO-ACEI-001", "Cambio de aceite y filtro", 0.75, 260, 550),
    ("MO-FREN-001", "Revision y servicio de frenos", 1.5, 520, 980),
    ("MO-SUSP-001", "Inspeccion de suspension y direccion", 1.25, 450, 850),
    ("MO-AIRE-001", "Diagnostico de aire acondicionado", 1, 400, 800),
]

DEMO_PARTS = [
    (
        "ESC-FIL-2020",
        "Filtro de aceite Ford Escape 2020",
        "FL-910S",
        "Motorcraft",
        "Unidad",
        *DEMO_VEHICLES[0],
        180,
        295,
    ),
    (
        "ESC-AIR-2020",
        "Filtro de aire Ford Escape 2020",
        "FA-1912",
        "Motorcraft",
        "Unidad",
        *DEMO_VEHICLES[0],
        310,
        495,
    ),
    (
        "ESC-PAD-2020",
        "Pastillas de freno delanteras Ford Escape 2020",
        "BRF-1554",
        "Motorcraft",
        "Juego",
        *DEMO_VEHICLES[0],
        1350,
        2190,
    ),
    (
        "F150-FIL-2020",
        "Filtro de aceite Ford F-150 2020",
        "FL-500S",
        "Motorcraft",
        "Unidad",
        *DEMO_VEHICLES[1],
        210,
        345,
    ),
    (
        "F150-AIR-2020",
        "Filtro de aire Ford F-150 2020",
        "FA-1883",
        "Motorcraft",
        "Unidad",
        *DEMO_VEHICLES[1],
        390,
        625,
    ),
    (
        "F150-PAD-2020",
        "Pastillas de freno delanteras Ford F-150 2020",
        "BRF-1522",
        "Motorcraft",
        "Juego",
        *DEMO_VEHICLES[1],
        1680,
        2690,
    ),
    (
        "CIV-FIL-2008",
        "Filtro de aceite Honda Civic 2008",
        "15400-PLM-A02",
        "Honda",
        "Unidad",
        *DEMO_VEHICLES[2],
        165,
        275,
    ),
    (
        "CIV-AIR-2008",
        "Filtro de aire Honda Civic 2008",
        "17220-RNA-A00",
        "Honda",
        "Unidad",
        *DEMO_VEHICLES[2],
        280,
        450,
    ),
    (
        "CIV-PAD-2008",
        "Pastillas de freno delanteras Honda Civic 2008",
        "45022-SNA-A00",
        "Honda",
        "Juego",
        *DEMO_VEHICLES[2],
        1050,
        1740,
    ),
]


def build_catalog_workbook(*, include_demo: bool = False) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instrucciones"
    instructions.append(["Plantilla de catálogo SmartDiag504"])
    instructions.append(
        ["No cambie los nombres de hojas ni encabezados. Una fila equivale a una compatibilidad."]
    )
    instructions.append(["Repita el mismo código para asociarlo con varios vehículos."])
    instructions.append(["El precio de costo incluye personal, local y demás costos internos."])
    instructions.append(["El precio de venta no puede ser menor al costo. Moneda: HNL."])
    instructions.column_dimensions["A"].width = 105
    instructions["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="C8102E")

    for sheet_name, headers in (("Mano de obra", LABOR_HEADERS), ("Repuestos", PART_HEADERS)):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="171717")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(
                14, min(30, len(header) + 4)
            )
        active_validation = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
        sheet.add_data_validation(active_validation)
        active_column = sheet.cell(1, headers.index("activo") + 1).column_letter
        active_validation.add(f"{active_column}2:{active_column}5000")

    if include_demo:
        labor_sheet = workbook["Mano de obra"]
        for code, description, hours, cost, sale in DEMO_LABOR:
            for make, model, year_from, year_to, engine in DEMO_VEHICLES:
                labor_sheet.append(
                    [
                        code,
                        description,
                        make,
                        model,
                        year_from,
                        year_to,
                        engine,
                        hours,
                        cost,
                        sale,
                        "SI",
                    ]
                )
        parts_sheet = workbook["Repuestos"]
        for (
            code,
            description,
            oem,
            brand,
            unit,
            make,
            model,
            year_from,
            year_to,
            engine,
            cost,
            sale,
        ) in DEMO_PARTS:
            parts_sheet.append(
                [
                    code,
                    description,
                    oem,
                    brand,
                    unit,
                    make,
                    model,
                    year_from,
                    year_to,
                    engine,
                    cost,
                    sale,
                    "SI",
                ]
            )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _text(value: Any) -> str:
    return str(value or "").strip()


def _decimal(
    value: Any, *, sheet: str, row: int, column: str, errors: list[ImportErrorRow]
) -> Decimal | None:
    try:
        result = Decimal(str(value))
        if result < 0:
            raise InvalidOperation
        return result
    except (InvalidOperation, ValueError, TypeError):
        errors.append(
            ImportErrorRow(sheet, row, column, "Debe ser un número mayor o igual a cero.")
        )
        return None


def _year(
    value: Any, *, sheet: str, row: int, column: str, errors: list[ImportErrorRow]
) -> int | None:
    try:
        result = int(value)
        if not 1900 <= result <= 2100:
            raise ValueError
        return result
    except (ValueError, TypeError):
        errors.append(ImportErrorRow(sheet, row, column, "Debe ser un año entre 1900 y 2100."))
        return None


def _parse_sheet(
    sheet: Any, headers: list[str], *, labor: bool, errors: list[ImportErrorRow]
) -> list[CatalogEntry]:
    actual = [_text(cell.value) for cell in sheet[1]]
    if actual != headers:
        errors.append(
            ImportErrorRow(
                sheet.title, 1, "encabezados", "Los encabezados no coinciden con la plantilla."
            )
        )
        return []
    grouped: dict[str, CatalogEntry] = {}
    invalid_codes: set[str] = set()
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        row = dict(zip(headers, values, strict=True))
        code = _text(row["codigo"]).upper()
        description = _text(row["descripcion"])
        row_error_count = len(errors)
        if not code:
            errors.append(
                ImportErrorRow(sheet.title, row_number, "codigo", "El código es obligatorio.")
            )
        if not description:
            errors.append(
                ImportErrorRow(
                    sheet.title, row_number, "descripcion", "La descripción es obligatoria."
                )
            )
        make = _text(row["marca_vehiculo"])
        model = _text(row["modelo_vehiculo"])
        if not make:
            errors.append(
                ImportErrorRow(
                    sheet.title, row_number, "marca_vehiculo", "La marca es obligatoria."
                )
            )
        if not model:
            errors.append(
                ImportErrorRow(
                    sheet.title, row_number, "modelo_vehiculo", "El modelo es obligatorio."
                )
            )
        year_from = _year(
            row["anio_desde"], sheet=sheet.title, row=row_number, column="anio_desde", errors=errors
        )
        year_to = _year(
            row["anio_hasta"], sheet=sheet.title, row=row_number, column="anio_hasta", errors=errors
        )
        if year_from and year_to and year_to < year_from:
            errors.append(
                ImportErrorRow(
                    sheet.title, row_number, "anio_hasta", "No puede ser menor que anio_desde."
                )
            )
        cost = _decimal(
            row["precio_costo_hnl"],
            sheet=sheet.title,
            row=row_number,
            column="precio_costo_hnl",
            errors=errors,
        )
        sale = _decimal(
            row["precio_venta_hnl"],
            sheet=sheet.title,
            row=row_number,
            column="precio_venta_hnl",
            errors=errors,
        )
        hours = None
        if labor:
            hours = _decimal(
                row["tiempo_horas"],
                sheet=sheet.title,
                row=row_number,
                column="tiempo_horas",
                errors=errors,
            )
            if hours == 0:
                errors.append(
                    ImportErrorRow(
                        sheet.title, row_number, "tiempo_horas", "Debe ser mayor que cero."
                    )
                )
        if cost is not None and sale is not None and sale < cost:
            errors.append(
                ImportErrorRow(
                    sheet.title,
                    row_number,
                    "precio_venta_hnl",
                    "El precio de venta no puede ser menor al costo.",
                )
            )
        if len(errors) != row_error_count:
            if code:
                invalid_codes.add(code)
            continue
        entry = grouped.get(code)
        if entry is None:
            entry = CatalogEntry(
                code=code,
                description=description,
                cost_price=cost or Decimal(0),
                sale_price=sale or Decimal(0),
                active=_text(row["activo"]).upper() not in {"NO", "0", "FALSE"},
                standard_hours=hours,
                oem_number=_text(row.get("numero_oem")) or None,
                brand=_text(row.get("marca_repuesto")) or None,
                unit=_text(row.get("unidad")) or None,
            )
            grouped[code] = entry
        elif (entry.description, entry.cost_price, entry.sale_price) != (description, cost, sale):
            errors.append(
                ImportErrorRow(
                    sheet.title,
                    row_number,
                    "codigo",
                    "Las filas repetidas deben conservar descripción y precios.",
                )
            )
            invalid_codes.add(code)
            continue
        fitment = VehicleFitment(
            make, model, year_from or 1900, year_to or 2100, _text(row["motor"]) or None
        )
        if fitment not in entry.fitments:
            entry.fitments.append(fitment)
    return [entry for code, entry in grouped.items() if code not in invalid_codes]


def parse_catalog_workbook(content: bytes) -> CatalogPreview:
    preview = CatalogPreview()
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception:
        preview.errors.append(
            ImportErrorRow("Archivo", 0, "archivo", "No es un archivo XLSX válido.")
        )
        return preview
    required = {"Mano de obra", "Repuestos"}
    missing = required.difference(workbook.sheetnames)
    for sheet_name in sorted(missing):
        preview.errors.append(ImportErrorRow(sheet_name, 0, "hoja", "Falta la hoja obligatoria."))
    if missing:
        return preview
    preview.labor = _parse_sheet(
        workbook["Mano de obra"], LABOR_HEADERS, labor=True, errors=preview.errors
    )
    preview.parts = _parse_sheet(
        workbook["Repuestos"], PART_HEADERS, labor=False, errors=preview.errors
    )
    return preview
